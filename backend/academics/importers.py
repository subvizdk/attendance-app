import datetime
from openpyxl import load_workbook
from django.utils.timezone import now
from django.db import transaction

from core.models import Branch
from .models import Student, Batch, StudentEnrollment


def parse_date(value):
    """
    Accepts:
      - Excel date/datetime
      - Python date/datetime
      - string 'YYYY-MM-DD'
    Returns:
      - datetime.date or None
    """
    if value is None or value == "":
        return None

    if isinstance(value, datetime.datetime):
        return value.date()

    if isinstance(value, datetime.date):
        return value

    try:
        return datetime.datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except Exception:
        return None


@transaction.atomic
def import_students_from_xlsx(file_obj):
    """
    Expected Excel columns (header row, row 1):

    Required:
      - branch_id (UUID)
      - admission_no
      - full_name

    Optional:
      - father_full_name
      - mother_full_name
      - email
      - phone
      - batch_id (int)          -> enroll student into this batch
      - start_date (YYYY-MM-DD) -> defaults to today if blank

    Behavior:
      - Student is upserted by (branch, admission_no)
      - If batch_id is provided, student is MOVED:
          - End any active enrollment (end_date set)
          - Create new active enrollment in the provided batch
      - Enforces: only 1 active enrollment per student (DB constraint)

    Returns:
      summary: {"created": int, "updated": int, "enrolled": int}
      errors:  [{"row": int, "error": str}, ...]
    """

    wb = load_workbook(file_obj)
    ws = wb.active

    # Read headers from row 1
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    required_cols = {"branch_id", "admission_no", "full_name"}
    missing = required_cols - set(headers)
    if missing:
        return (
            {"created": 0, "updated": 0, "enrolled": 0},
            [{"row": 1, "error": f"Missing required columns: {', '.join(sorted(missing))}"}],
        )

    # Map header name -> index in row list
    idx = {h: i for i, h in enumerate(headers)}

    created = 0
    updated = 0
    enrolled = 0
    errors = []

    # Iterate data rows starting row 2
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

        try:
            branch_id = row[idx["branch_id"]]
            admission_no = row[idx["admission_no"]]
            full_name = row[idx["full_name"]]

            # Optional fields
            father_name = row[idx["father_full_name"]] if "father_full_name" in idx else ""
            mother_name = row[idx["mother_full_name"]] if "mother_full_name" in idx else ""
            email = row[idx["email"]] if "email" in idx else ""
            phone = row[idx["phone"]] if "phone" in idx else ""

            batch_id = row[idx["batch_id"]] if "batch_id" in idx else None
            start_date_val = row[idx["start_date"]] if "start_date" in idx else None

            # Validate required
            if not branch_id or not admission_no or not full_name:
                errors.append({"row": r, "error": "branch_id, admission_no, and full_name are required"})
                continue

            # Validate branch
            try:
                branch = Branch.objects.get(id=str(branch_id).strip())
            except Branch.DoesNotExist:
                errors.append({"row": r, "error": f"Branch not found for branch_id={branch_id}"})
                continue

            # Normalize strings
            admission_no = str(admission_no).strip()
            full_name = str(full_name).strip()

            father_name = str(father_name).strip() if father_name else ""
            mother_name = str(mother_name).strip() if mother_name else ""
            email = str(email).strip() if email else ""
            phone = str(phone).strip() if phone else ""

            # Upsert by (branch, admission_no)
            student, was_created = Student.objects.update_or_create(
                branch=branch,
                admission_no=admission_no,
                defaults={
                    "full_name": full_name,
                    "father_full_name": father_name,
                    "mother_full_name": mother_name,
                    "email": email,
                    "phone": phone,
                },
            )

            if was_created:
                created += 1
            else:
                updated += 1

            # Optional enrollment into batch (MOVE)
            if batch_id not in (None, ""):
                try:
                    batch = Batch.objects.get(id=int(batch_id))
                except Exception:
                    errors.append({"row": r, "error": f"Invalid batch_id={batch_id}"})
                    continue

                # Ensure batch belongs to same branch
                if batch.branch_id != branch.id:
                    errors.append({"row": r, "error": f"batch_id={batch_id} is not in the same branch as branch_id={branch.id}"})
                    continue

                start_date = parse_date(start_date_val) or now().date()

                # End any active enrollment (student can only be in one batch at a time)
                StudentEnrollment.objects.filter(student=student, end_date__isnull=True).update(end_date=now().date())

                StudentEnrollment.objects.create(
                    student=student,
                    batch=batch,
                    start_date=start_date,
                    end_date=None,
                )
                enrolled += 1

        except Exception as e:
            errors.append({"row": r, "error": str(e)})

    summary = {"created": created, "updated": updated, "enrolled": enrolled}
    return summary, errors